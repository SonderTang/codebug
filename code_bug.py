import calendar
import os
import subprocess
import time
import openpyxl
import sys
from typing import List
from alibabacloud_devops20210625.client import Client as devops20210625Client
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_devops20210625 import models as devops_20210625_models
from alibabacloud_tea_util import models as util_models
from alibabacloud_tea_util.client import Client as UtilClient
import flask
import pymysql
from flask import Flask
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font


service = Flask(__name__)

@service.route('/thousand_line_code_bug_rate', methods=['post'])
def thousand_line_code_bug_rate():
    # 计算每月千行代码bug率
    mouth = flask.request.values.get('mouth')  # 月份
    year = flask.request.values.get('year')  # 年份
    query_bug_data() # 查询bug数据
    query_code_date(year, mouth) # 查询代码数量
    query_code_bug_date(year, mouth) # 计算千行代码bug率
    c = file_thousand_line_code_bug_rate(year, mouth)
    return c



def file_thousand_line_code_bug_rate(year,mouth):
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    db4 = db3

    cur4 = db4.cursor()

    # file_path2 = "/Users/biden/code/yangteng/质量管理专项/缺陷爬虫/文件/千行代码bug率.xlsx"
    file_path2 = "/home/jumper/codebug/pyscript/千行代码bug率.xlsx"
    table_write = openpyxl.load_workbook(file_path2)
    print('################################!!!!!!')
    today = year + '-' + mouth
    today1 = "'"+today+"'"
    print('today1'+ today1)
    cur3.execute('select * from thousand_line_code_bug_rate2 a where a.month = %s;'%today1)
    result_a = cur3.fetchall()
    len_a = len(result_a)
    print('len_a'+str(len_a))
    table1 = table_write.worksheets[0]
    table2 = table_write.worksheets[1]
    table3 = table_write.worksheets[2]
    table4 = table_write.worksheets[3]

    print('table')
    print(result_a)

    for i in range(len_a):
        a = i + 2
        id = i+1
        print('a', a)
        print('id', id)
        name = result_a[i][1]
        en_name = result_a[i][2]
        month = result_a[i][3]
        num = result_a[i][4]
        number = result_a[i][5]
        rate = result_a[i][6]
        rate1 = str(round(rate*1000,4))+'‰'

        table1.cell(row=a, column=1).value = id
        table1.cell(row=a, column=2).value = name
        table1.cell(row=a, column=3).value = en_name
        table1.cell(row=a, column=4).value = month
        table1.cell(row=a, column=5).value = num
        table1.cell(row=a, column=6).value = number
        table1.cell(row=a, column=7).value = rate1
        print('结束', id)

    cur3.execute('select * from thousand_line_code_bug_rate a where a.mouth = %s and a.number != 0;'%today1)
    result2 = cur3.fetchall()
    len_result2 = len(result2)
    print('len_result2:'+str(len_result2))
    for n in range(len_result2):
        b = n + 2
        name1 = result2[n][1]
        project1 = result2[n][2]
        mouth1 = result2[n][3]
        number1 = result2[n][4]

        table2.cell(row=b, column=1).value = name1
        table2.cell(row=b, column=2).value = project1
        table2.cell(row=b, column=3).value = mouth1
        table2.cell(row=b, column=4).value = number1

    cur4.execute('select * from project_name a;')
    result3 = cur4.fetchall()
    len_result3 = len(result3)
    print('len_result3:' + str(len_result3))
    for m in range(len_result3):
        c = m + 2
        project_name = result3[m][1]
        project = result3[m][2]
        branch = result3[m][6]

        table3.cell(row=c, column=1).value = project_name
        table3.cell(row=c, column=2).value = project
        table3.cell(row=c, column=3).value = branch

    cur4.execute('select * from person a where a.code =1;')
    result4 = cur4.fetchall()
    len_result4 = len(result4)
    print('len_result4:' + str(len_result4))
    for l in range(len_result4):
        k = l +2
        person = result4[l][1]
        en_person = result4[l][2]

        table4.cell(row=k, column=1).value = person
        table4.cell(row=k, column=2).value = en_person

    name_1 = today + '千行代码bug率.xlsx'
    table_write.save('/home/jumper/codebug/pyscript/' + name_1)
    file_name = '千行代码bug率/' + name_1
    file_path = '/Users/biden/code/yangteng/质量管理专项/缺陷爬虫/文件/' + name_1
    return  name_1


def str_to_timestamp_code(time_data):
    #string类型转时间戳
    temp1 = time.strptime(time_data,"%Y-%m-%d")
    temp = time.mktime(temp1)

    return temp



def query_code_bug_date(year,mouth):
    print('year:'+year)
    print('mouth:'+mouth)
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    db4 = db3
    # db4 = pymysql.connect(
    #     host='101.43.40.158',
    #     port=3306,
    #     user='yunxiaodata',
    #     password='filang123',
    #     charset='utf8',
    #     database='yunxiaodata')
    cur4 = db4.cursor()

    #初始化该月的thousand_line_code_bug_rate2数据
    cur4.execute('select * from person a where a.code = 1;')
    name_result = cur4.fetchall()
    print('name_result: ')
    print('name_result: ' + str(name_result))
    len_name = len(name_result)
    code = year + '-' + mouth
    for i in range(len_name):
        name = name_result[i][1]
        en_name = name_result[i][2]
        cur3.execute('select * from thousand_line_code_bug_rate2 a order by a.id desc limit 1;')
        id1 = cur3.fetchone()
        if id1 is None:
            id = 1
        else:
            id = id1[0]+1

        #判断姓名月份是否存在thousand_line_code_bug_rate2，存在不处理，不存在新增
        cur3.execute('select * from thousand_line_code_bug_rate2 a where a.name = %s and a.month =%s;',(name,code))
        result1 = cur3.fetchone()
        if result1 is None:
            cur3.execute('INSERT INTO thousand_line_code_bug_rate2(id,name,en_name,month)VALUES(%s,%s,%s,%s);',(id,name,en_name,code))
            db3.commit()

    sql = 'select * from work_week_day a where a.mouth_day like "%'+code+'%" order by id desc limit 1;'
    cur3.execute(sql)

    ids =cur3.fetchone()[0]
    id = int(ids)+1
    print('id:' + str(id))
    cur3.execute('select a.mouth_day from work_week_day a where a.id = %s;'%id)
    mouth_day_a = cur3.fetchone()
    print('mouth_day_a'+ str(mouth_day_a))
    mouth_day_b = mouth_day_a[0]


    start_time = code +'-01'
    end_time = mouth_day_b

    start_time1 = str_to_timestamp_code(start_time)
    end_time1 = str_to_timestamp_code(end_time)
    start_time2 = int(start_time1)
    end_time2 = int(end_time1)

    #查询对应开发的bug数
    code1 = "'"+code+"'"
    cur3.execute('select * from thousand_line_code_bug_rate2 a where a.month = %s;'%code1)
    result1 = cur3.fetchall()
    len_result1 = len(result1)
    for u in range(len_result1):
        name = result1[u][1]
        assignedTo = "'"+name+"'"
        id2 = result1[u][0]

        cur3.execute('select count(1) from bug_data a where a.gmtCreate>%s and a.gmtCreate<%s and a.assignedTo = %s and a.status in ("关闭（已验证）", "再次打开", "处理中", "已修复", "已关闭", "延期处理", "待确认", "暂不修复", "设计如此", "转需求");',(start_time2,end_time2,name))
        result = cur3.fetchone()
        result_1 = result[0]

        cur3.execute('update thousand_line_code_bug_rate2 a set a.num =%s where a.id = %s;',(result_1,id2))
        db3.commit()


    #查询对应开发代码行数
    cur3.execute('select * from thousand_line_code_bug_rate2 a where a.month = %s;'%code1)
    result2 = cur3.fetchall()
    len_result2 = len(result2)

    for t in range(len_result2):
        en_name = result2[t][1]
        en_name1 = "'"+en_name+"'"
        id3 = result2[t][0]
        cur3.execute('select sum(number) from thousand_line_code_bug_rate a where a.name = %s and a.mouth = %s;',(en_name,code))
        code_result = cur3.fetchone()[0]
        if code_result is None:
            code_result = 0

        cur3.execute('update thousand_line_code_bug_rate2 a set a.number = %s where a.id = %s;',(code_result,id3))
        db3.commit()


    cur3.execute('select * from thousand_line_code_bug_rate2 a where a.month = %s;'%code1)
    result3 = cur3.fetchall()
    len_result3 = len(result3)
    for r in range(len_result3):
        id4 = result3[r][0]
        num1 = result3[r][4]
        number1 = result3[r][5]
        if number1 ==0:
            rate = 0
        else:
            rate = float(num1/number1)
        cur3.execute('update thousand_line_code_bug_rate2 a set a.rate = %s where a.id = %s;',(rate,id4))
        db3.commit()


    cur3.execute('select count(1) from thousand_line_code_bug_rate2;')
    count = cur3.fetchone()[0]

    cur3.close()
    db3.close()
    cur4.close()
    # db4.close()

    return count



# spacename 表，并返回两个列表：spaceName_list 和 space_list，它们分别包含表中的 spaceName 和 space 字段的值
def header_code_return():
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    cur3.execute('select * from spacename;')
    result = cur3.fetchall()
    len_result = len(result)
    spaceName_list =[]
    space_list =[]
    for i in range(len_result):
        result2 = result[i]

        spaceName = result2[1]
        space = result2[2]
        spaceName_list.append(spaceName)
        space_list.append(space)
    return spaceName_list,space_list


class bug_query:
    def __init__(self):
        pass

    @staticmethod
    def create_client(
        access_key_id: str,
        access_key_secret: str,
    ) -> devops20210625Client:
        """
        使用AK&SK初始化账号Client
        @param access_key_id:
        @param access_key_secret:
        @return: Client
        @throws Exception
        """
        config = open_api_models.Config(
            # 必填，您的 AccessKey ID,
            access_key_id='LTAI5tDruhW5L33cPNUpLTLE',
            # 必填，您的 AccessKey Secret,
            access_key_secret='uy7sH7ZJaFqiLpa85d2fXterJnMEWu'
        )
        # Endpoint 请参考 https://api.aliyun.com/product/devops
        config.endpoint = f'devops.cn-hangzhou.aliyuncs.com'
        return devops20210625Client(config)

    @staticmethod
    def main(next_token,space_identifier):
        # 请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID 和 ALIBABA_CLOUD_ACCESS_KEY_SECRET。
        # 工程代码泄露可能会导致 AccessKey 泄露，并威胁账号下所有资源的安全性。以下代码示例使用环境变量获取 AccessKey 的方式进行调用，仅供参考，建议使用更安全的 STS 方式，更多鉴权访问方式请参见：https://help.aliyun.com/document_detail/378659.html
        # client = bug_query.create_client(os.environ['ALIBABA_CLOUD_ACCESS_KEY_ID'], os.environ['ALIBABA_CLOUD_ACCESS_KEY_SECRET'])
        client = bug_query.create_client('LTAI5tDruhW5L33cPNUpLTLE', 'uy7sH7ZJaFqiLpa85d2fXterJnMEWu')
        list_workitems_request = devops_20210625_models.ListWorkitemsRequest(
            space_type='Project',
            space_identifier=space_identifier,
            category='Bug',
            next_token=next_token,
            max_results = '200',
            conditions='{"conditionGroups": [[{"fieldIdentifier": "gmtCreate","operator": "MORE_THAN_AND_EQUAL","value": ["2025-03-01 00:00:00"]}]]}',
        )
        runtime = util_models.RuntimeOptions()
        headers = {}
        try:
            # 复制代码运行请自行打印 API 的返回值
            ListWorkitemsResponse = client.list_workitems_with_options('5ffffa7f1e45db3c1cc27185', list_workitems_request, headers, runtime)
            # a =  json.dumps(ListWorkitemsResponse)
            return ListWorkitemsResponse
        except Exception as error:
            # 如有需要，请打印 error
            UtilClient.assert_as_string(error)
            print(UtilClient.assert_as_string(error))

# class query_reOpen:
#     def __init__(self):
#         pass
#     @staticmethod
#     def create_client(
#         access_key_id: str,
#         access_key_secrect: str
#     ) -> devops20210625Client:


class user_query:
    def __init__(self):
        pass

    @staticmethod
    def create_client(
        access_key_id: str,
        access_key_secret: str,
    ) -> devops20210625Client:
        """
        使用AK&SK初始化账号Client
        @param access_key_id:
        @param access_key_secret:
        @return: Client
        @throws Exception
        """
        config = open_api_models.Config(
            # 必填，您的 AccessKey ID,
            access_key_id=access_key_id,
            # 必填，您的 AccessKey Secret,
            access_key_secret=access_key_secret
        )
        # Endpoint 请参考 https://api.aliyun.com/product/devops
        config.endpoint = f'devops.cn-hangzhou.aliyuncs.com'
        return devops20210625Client(config)

    @staticmethod
    def main(organizationid):
        # 请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID 和 ALIBABA_CLOUD_ACCESS_KEY_SECRET。
        # 工程代码泄露可能会导致 AccessKey 泄露，并威胁账号下所有资源的安全性。以下代码示例使用环境变量获取 AccessKey 的方式进行调用，仅供参考，建议使用更安全的 STS 方式，更多鉴权访问方式请参见：https://help.aliyun.com/document_detail/378659.html
        client = user_query.create_client('LTAI5tDruhW5L33cPNUpLTLE', 'uy7sH7ZJaFqiLpa85d2fXterJnMEWu')
        # client = user_query.create_client(os.environ['ALIBABA_CLOUD_ACCESS_KEY_ID'],
        #                                   os.environ['ALIBABA_CLOUD_ACCESS_KEY_SECRET'])

        list_project_members_request = devops_20210625_models.ListProjectMembersRequest(
            target_type='Space'
        )
        runtime = util_models.RuntimeOptions()
        headers = {}
        try:
            # 复制代码运行请自行打印 API 的返回值
            result = client.list_project_members_with_options('5ffffa7f1e45db3c1cc27185', organizationid, list_project_members_request, headers, runtime)
            a = result
            return result
        except Exception as error:
            # 如有需要，请打印 error
            UtilClient.assert_as_string(error.message)

    @staticmethod
    async def main_async(
        args: List[str],
    ) -> None:
        # 请确保代码运行环境设置了环境变量 ALIBABA_CLOUD_ACCESS_KEY_ID 和 ALIBABA_CLOUD_ACCESS_KEY_SECRET。
        # 工程代码泄露可能会导致 AccessKey 泄露，并威胁账号下所有资源的安全性。以下代码示例使用环境变量获取 AccessKey 的方式进行调用，仅供参考，建议使用更安全的 STS 方式，更多鉴权访问方式请参见：https://help.aliyun.com/document_detail/378659.html
        client = user_query.create_client(os.environ['ALIBABA_CLOUD_ACCESS_KEY_ID'], os.environ['ALIBABA_CLOUD_ACCESS_KEY_SECRET'])
        list_project_members_request = devops_20210625_models.ListProjectMembersRequest(
            target_type='Space'
        )
        runtime = util_models.RuntimeOptions()
        headers = {}
        try:
            # 复制代码运行请自行打印 API 的返回值
            await client.list_project_members_with_options_async('5ffffa7f1e45db3c1cc27185', '2f47d6d1e8613e642d7abe6d99', list_project_members_request, headers, runtime)
        except Exception as error:
            # 如有需要，请打印 error
            UtilClient.assert_as_string(error.message)

# class reOpen_query:
#     def __init__(self):
#         pass
    

def query_bug_data():
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    cur3.execute('delete from bug_data;')
    db3.commit()

    header_code = header_code_return()[0]
    print('头部code'+ str(header_code))
    len_header_code = len(header_code)
    for p in range(len_header_code):
        next_token1 = ''
        header_code_result = header_code[p]
        first_request = bug_query.main(next_token1,header_code_result)
        total_count = first_request.body.total_count
        count = int(total_count) // 200+1


        project = ''
        next_token = ''

        for i in range(count):
            request_data = bug_query.main(next_token,header_code[p])
            next_token = request_data.body.next_token
            max_results = len(request_data.body.workitems)
            # print('max_results' + str(max_results))
            for y in range(max_results):
                assignedTo = request_data.body.workitems[y].assigned_to
                categoryIdentifier = request_data.body.workitems[y].category_identifier
                creator = request_data.body.workitems[y].creator
                document = request_data.body.workitems[y].document
                gmtCreate = request_data.body.workitems[y].gmt_create
                gmtModified = request_data.body.workitems[y].gmt_modified
                identifier = request_data.body.workitems[y].identifier
                logicalStatus = request_data.body.workitems[y].logical_status
                modifier = request_data.body.workitems[y].modifier
                parentIdentifier = request_data.body.workitems[y].parent_identifier
                serialNumber = request_data.body.workitems[y].serial_number
                spaceIdentifier = request_data.body.workitems[y].space_identifier
                spaceName = request_data.body.workitems[y].space_name
                spaceType = request_data.body.workitems[y].space_type
                sprintIdentifier = request_data.body.workitems[y].sprint_identifier
                status = request_data.body.workitems[y].status
                statusIdentifier = request_data.body.workitems[y].status_identifier
                statusStageIdentifier = request_data.body.workitems[y].status_stage_identifier
                subject = request_data.body.workitems[y].subject
                workitemTypeIdentifier = request_data.body.workitems[y].workitem_type_identifier
                detailLink = 'https://devops.aliyun.com/projex/project/' + spaceIdentifier + '/bug/' + identifier
                cur3.execute('select a.id from bug_data a order by id desc limit 1;')
                ids = cur3.fetchone()
                if ids is None:
                    id = 1
                else:
                    id = ids[0] +1
                cur3.execute('insert into bug_data (id,assignedTo,categoryIdentifier,creator,document,gmtCreate,gmtModified,identifier,logicalStatus,modifier,parentIdentifier,serialNumber,spaceIdentifier,spaceName,spaceType,sprintIdentifier,status,statusIdentifier,statusStageIdentifier,subject,workitemTypeIdentifier,detail_link) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);',(id,assignedTo,categoryIdentifier,creator,document,gmtCreate,gmtModified,identifier,logicalStatus,modifier,parentIdentifier,serialNumber,spaceIdentifier,spaceName,spaceType,sprintIdentifier,status,statusIdentifier,statusStageIdentifier,subject,workitemTypeIdentifier,detailLink))
                db3.commit()
                print('query_bug_data '+ str(p)+ ' '+str(project) +' ' + str(next_token)+ ' '+str(i*200+y))


    cur3.execute('select count(1) from bug_data a ;')
    count = cur3.fetchone()[0]
    for m in range(count):
        n = m+1
        cur3.execute('select a.creator from bug_data a where a.id = %s'%n)
        creator_id = cur3.fetchone()[0]
        cur3.execute('select a.realName from user_data a where a.identifier = %s;'%creator_id)
        realName_result = cur3.fetchone()
        if realName_result is not None:
                realName = realName_result[0]
                cur3.execute('update bug_data a set a.creator = %s where a.id =%s;',(realName,n))
                db3.commit()
        cur3.execute('select a.gmtCreate,a.gmtModified from bug_data a where a.id = %s'%n)
        date_result = cur3.fetchone()
        gmtCreate_1 = int(date_result[0])/1000
        gmtModified_1 = int(date_result[1])/1000
        gmtCreate_2 = time.localtime(gmtCreate_1)
        gmtModified_2 = time.localtime(gmtModified_1)
        gmtCreate = time.strftime("%Y-%m-%d %H:%M:%S",gmtCreate_2)
        gmtModified = time.strftime("%Y-%m-%d %H:%M:%S",gmtModified_2)
        cur3.execute('update bug_data a set a.gmtCreate= %s,a.gmtModified=%s where a.id = %s;',(gmtCreate_1,gmtModified_1,n))
        db3.commit()
        cur3.execute('select a.modifier from bug_data a where a.id = %s'%n)
        creator_id = cur3.fetchone()[0]
        if creator_id is not None:
            cur3.execute('select a.realName from user_data a where a.identifier = %s;'%creator_id)
            realName_result = cur3.fetchone()
            if realName_result is not None:
                    realName = realName_result[0]
                    cur3.execute('update bug_data a set a.modifier = %s where a.id =%s;',(realName,n))
                    db3.commit()

    # 更新bug_data的assignedTo字段值，负责人
    cur3.execute('select * from bug_data a;')
    bug_result = cur3.fetchall()
    len_bug_result = len(bug_result)
    for o in range(len_bug_result):
        assignedTo = bug_result[o][4]
        id_bug = bug_result[o][0]
        cur3.execute('select a.realName from user_data a where a.identifier = %s;' % assignedTo)
        realName = cur3.fetchone()
        if realName is None:
            realName_result = '查无此人'
        else:
            realName_result = realName[0]
        cur3.execute('update bug_data a set a.assignedTo = %s where a.id =%s;', (realName_result, id_bug))
        db3.commit()

    cur3.close()
    db3.close()



def number():
    # 数据库链接
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()

    #开发数据统计
    cur3.execute("SELECT COUNT(1) FROM person a WHERE a.code =1;")
    develop_number = cur3.fetchone()[0]

    #项目数据统计
    cur3.execute("SELECT COUNT(1) FROM project_name a where a.is_check =1;")
    project_number = cur3.fetchone()[0]

    cur3.close()
    db3.close()
    return develop_number,project_number


def insert_user_data():
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    cur3.execute('delete from user_data;')
    db3.commit()
    cur3.execute(
        'insert into user_data (id,identifier,realName) values(%s,%s,%s);',
        (1, '10000', '自动化规则'))
    db3.commit()

    header_code = header_code_return()[0]
    len_header_code = len(header_code)
    for p in range(len_header_code):
        first_user_query = user_query.main(header_code[p])
        cur3.execute('select a.id from user_data a order by id desc limit 1;')
        ids2 = cur3.fetchone()
        if ids2 is None:
            id1 = 1
        else:
            id1 = ids2[0] + 1
        a = len(first_user_query.body.members)
        for l in range(a):
            avatar = first_user_query.body.members[l].avatar
            displayName = first_user_query.body.members[l].display_name
            division = first_user_query.body.members[l].division.identifier
            identifier = first_user_query.body.members[l].identifier
            nickName = first_user_query.body.members[l].nick_name
            organizationUserInfo = first_user_query.body.members[l].organization_user_info.organization_identifier
            realName = first_user_query.body.members[l].real_name
            roleName = first_user_query.body.members[l].role_name
            tbRoleId = first_user_query.body.members[l].tb_role_id
            cur3.execute('select * from user_data a where a.identifier = %s'%identifier)
            is_identifier = cur3.fetchone()
            if is_identifier is None:
                cur3.execute('insert into user_data (id,avatar,displayName,division,identifier,nickName,organizationUserInfo,realName) values(%s,%s,%s,%s,%s,%s,%s,%s);',(id1,avatar,displayName,division,identifier,nickName,organizationUserInfo,realName))
                db3.commit()
            id1 = id1 + 1



def date_data4(year1,mouth1):
    year2 = int(year1)
    mouth2 = int(mouth1)
    first_day, num_days = calendar.monthrange(year2, mouth2)
    # print(1)
    start_day = year1+'-'+mouth1+'-01'
    end_day = year1+'-'+mouth1+'-'+str(num_days)
    return start_day,end_day



def project_name(c):
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    a = c + 1
    cur3.execute("SELECT a.project_name FROM project_name a where a.id = %s" % a)
    project_name = cur3.fetchone()[0]

    cur3.close()
    db3.close()
    return project_name



def code_config(project, v, ids):
    code_path = ''
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()
    cur3.execute("SELECT a.code_path FROM project_name a WHERE a.id = '%s'"%ids)
    code_path = cur3.fetchone()[0]
    cur3.execute("SELECT a.en_person FROM person a WHERE a.code = 1 ORDER BY id ASC LIMIT %s,1"%v)
    name = cur3.fetchone()[0]
    cur3.execute("SELECT a.person FROM person a WHERE a.code = 1 ORDER BY id ASC LIMIT %s,1" % v)
    name1 = cur3.fetchone()[0]

    cur3.close()
    db3.close()

    return name,code_path,name1



def query_code_date(year,mouth):
    db3 = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur3 = db3.cursor()

    number_result = number() # 统计开发者数量和项目数量

    mouth1 = mouth.split(",")
    len_mouth = len(mouth1)
    cur3.execute('select * from thousand_line_code_bug_rate a order by id desc limit 1;')
    result4 = cur3.fetchone()
    if result4 is None:
        id = 1
    else:
        id = result4[0]+1


    for u in range(len_mouth):
        mouth2 = mouth1[u]
        date_time = date_data4(year,mouth)
        start_time = date_time[0]
        end_time = date_time[1]
        month = year+'-'+mouth2
        month_a = "'"+month+"'"

        # cur3.execute('delete from thousand_line_code_bug_rate a where a.mouth = %s;' % month_a)
        cur3.execute('delete from thousand_line_code_bug_rate;')
        db3.commit()
        # cur3.execute('delete from thousand_line_code_bug_rate2 a where a.month = %s;' % month_a)
        cur3.execute('delete from thousand_line_code_bug_rate2;')
        db3.commit()
        print('项目数量', number_result[1])
        for c in range(number_result[1]):#项目数量

            for v in range(number_result[0]):#开发数量
                project = project_name(c)
                ids = c + 1
                name_auther = code_config(project, v, ids)[0]
                name_auther1 = code_config(project, v, ids)[2]

                # wms 项目兼容,继亮和杨静的提交名字在wms项目中不一样
                # if ids == 13 and name_auther == "Chris Qiu":
                #     name_auther = "Chris"
                # if ids == 13 and name_auther == "Jane Yang":
                #     name_auther = "Jane"
                # if ids == 8 and name_auther == "kermit jiang":
                #     name_auther = "jiangqm"

                code_path = code_config(project, v, ids)[1]
                # print('code_path'+ code_path)
                os.chdir(code_path)
                # git log --author=xxx --after='2024-08-01' --before='2024-08-31' --no-merges --shortstat --pretty=startLog%an
                input_name1 ="git log --author=%s" % str('"'+name_auther+'"')
                input_name2 = "--after='%s'" % str(start_time)
                input_name3 = "--before='%s'" % str(end_time)
                input_name4 = "--no-merges" # 只显示非合并记录
                # input_name7 = "--grep='^Merge branch 'release'' -v"
                input_name5 = "--shortstat" # 它会为每个提交的日志条目附加一个简短的统计信息，显示该提交中修改了多少文件，以及在这些文件中添加了多少行和删除了多少行。
                input_name6 = "--pretty=startLog%an"
                input_name7 = "-- ':!package-lock.json'"
                input_name = input_name1 + ' ' + input_name2 + ' ' + input_name3 + ' ' + input_name4 + ' ' + input_name5 + ' ' + input_name6 + ' ' + input_name7
                print('gitlab日志打印' + input_name)
                repo21 = subprocess.Popen(input_name,stdin=None,stdout=subprocess.PIPE,stderr=subprocess.PIPE,shell=True)
                repo2 = str(repo21.communicate()[0])
                repo2 = repo2.replace('\\n', '')
                repo2 = repo2.split('startLog')
                repo2 = [x.strip() for x in repo2 if x.strip() != ''] #删除字符串前后空格
                b = 0
                # print(repo2)
                # print('lenrepo2', len(repo2))
                for i in range(len(repo2)):
                    if 'insertion' in repo2[i]:
                        a = int(repo2[i].split(',')[1].split(' ins')[0])
                        b = b + a

                #查询thousand_line_code_bug_rate是否存在该人该月份该服务的记录，存在相加，不存在新增
                cur3.execute('select * from thousand_line_code_bug_rate a where a.name = %s and a.project = %s and a.mouth = %s;',(name_auther1,project,month))
                result5 = cur3.fetchone()
                if result5 is None:
                    cur3.execute('INSERT INTO thousand_line_code_bug_rate(id,name,mouth,number,project)VALUES(%s,%s,%s,%s,%s);',(id,name_auther1,month,b,project))
                    db3.commit()

                else:
                    id = result5[0]
                    number1 = int(result5[4])
                    b = b + number1
                    cur3.execute('update thousand_line_code_bug_rate a set a.number = %s where a.id = %s;',(b,id))
                    db3.commit()


                id = id +1
                print('项目：' + project + '    ' + '开发 人员' + name_auther + '    ' + str(b))

    cur3.execute('select count(1) from thousand_line_code_bug_rate;')
    count = cur3.fetchone()[0]
    cur3.close()
    db3.close()
    return str(count)

def file_bug_cause():
    db_yx = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur_yx = db_yx.cursor()
    db_bug = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur_bug = db_bug.cursor()

    file_path = "/home/jumper/codebug/pyscript/缺陷原因.xlsx"
    table_write = openpyxl.load_workbook(file_path)

    cur_yx.execute('select * from bug_cause')
    result_a = cur_yx.fetchall()
    len_a = len(result_a)
    table = table_write.worksheets[0]

    for i in range(len_a):
        a = i + 2
        id = i+1
        title = result_a[i][1]
        reason = result_a[i][2]
        table.cell(row=a, column=1).value = title
        table.cell(row=a, column=2).value = reason
        # table.cell(row=a, column=3).value = reason

    name_1 = 'bug缺陷原因.xlsx'
    table_write.save('/home/jumper/codebug/pyscript/' + name_1)
    return name_1

# 分项目导出bug数据
def create_mrp_bug_cause_table():
    db_yx = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur_yx = db_yx.cursor()
    db_bug = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')

    project_list = ['独立站项目组', 'MRP', 'ERP', '扬腾仓储', 'OMS']
    len_pro = len(project_list)
    for j in range(len_pro):
        cur_bug = db_bug.cursor()
        file_path = "/home/jumper/codebug/pyscript/缺陷复盘表格.xlsx"
        table_write = openpyxl.load_workbook(file_path)
        current_pro = project_list[j]

        cur_yx.execute("select * from front_bug_data where spaceName = ?", (current_pro,))
        result_a = cur_yx.fetchall()
        len_a = len(result_a)
        table = table_write.worksheets[0]

        for i in range(len_a):
            a = i + 2
            id = i + 1
            title = result_a[i][1]
            reason = result_a[i][6]
            assignedTo = result_a[i][4]
            link = 'https://devops.aliyun.com/projex/project/' + result_a[i][14] + '/bug/' + result_a[i][9]

            table.cell(row=a, column=2).value = title
            table.cell(row=a, column=3).value = reason
            table.cell(row=a, column=4).value = assignedTo
            table.cell(row=a, column=5).value = '点击跳转'
            font = Font(color='589DF6')  # 蓝色
            table.cell(row=a, column=5).font = font
            table.cell(row=a, column=5).hyperlink = link

        name_1 = current_pro+'一月前端bug复盘表格.xlsx'
        table_write.save('/home/jumper/codebug/pyscript/' + name_1)
    return 'a'

# 分项目导出bug数据
def create_front_bug_cause_table():
    db_yx = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')
    cur_yx = db_yx.cursor()
    db_bug = pymysql.connect(
        host='101.43.40.158',
        port=3306,
        user='yunxiaodata',
        password='filang123',
        charset='utf8',
        database='yunxiaodata')

    project_list = ['独立站项目组', 'MRP', 'ERP', '扬腾仓储', 'OMS']
    cur_bug = db_bug.cursor()
    file_path = "/home/jumper/codebug/pyscript/缺陷复盘表格.xlsx"
    table_write = openpyxl.load_workbook(file_path)

    cur_yx.execute("select * from front_bug_data")
    result_a = cur_yx.fetchall()
    len_a = len(result_a)
    table = table_write.worksheets[0]

    for i in range(len_a):
        a = i + 2
        id = i + 1
        title = result_a[i][1]
        reason = result_a[i][6]
        assignedTo = result_a[i][4]
        link = 'https://devops.aliyun.com/projex/project/' + result_a[i][14] + '/bug/' + result_a[i][9]
        projectName = result_a[i][15]

        table.cell(row=a, column=2).value = title
        table.cell(row=a, column=3).value = reason
        table.cell(row=a, column=4).value = assignedTo
        table.cell(row=a, column=5).value = projectName
        table.cell(row=a, column=6).value = link
        # table.cell(row=a, column=5).value = '点击跳转'
        # font = Font(color='589DF6')  # 蓝色
        # table.cell(row=a, column=5).font = font
        # table.cell(row=a, column=5).hyperlink = link

    name_1 = '二月前端bug复盘表格.xlsx'
    table_write.save('/home/jumper/codebug/pyscript/' + name_1)

    return 'a'


if __name__ == '__main__':

    # insert_user_data()
    service.run(port=5001, debug=True, host='0.0.0.0')
    # file_bug_reply()
    # file_bug_cause()
    # file_thousand_line_code_bug_rate('2024', '08')
    # create_mrp_bug_cause_table()
    # create_front_bug_cause_table()